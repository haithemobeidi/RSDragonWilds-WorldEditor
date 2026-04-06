#!/usr/bin/env python3
"""
Import the Ashenfall Completionist Log XLSX and convert it to JSON catalog files.

The original XLSX is by Accurious (RSDW Discord community). This script extracts
the Quest Journal and Collection Log into JSON files the editor can use as a
reference database.

Note: The XLSX uses community IDs (MQ01, L001 etc.), NOT the game's internal
GUIDs. So this catalog can only be used for browsing/reference UNTIL we have
a separate GUID→catalog-id mapping (Phase 2 — requires .pak extraction).

Usage:
    python scripts/import_catalog.py [path/to/xlsx]

Default XLSX path: ~/Downloads/Copy of Ashenfall's Completionist Log.xlsx
Outputs:
    data/quests.json
    data/items.json
    data/catalog_meta.json   (versioning + source info)
"""
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

DEFAULT_XLSX = os.path.expanduser(
    "~/Downloads/Copy of Ashenfall's Completionist Log.xlsx"
)
# WSL fallback
if not os.path.exists(DEFAULT_XLSX):
    DEFAULT_XLSX = "~/Downloads/Copy of Ashenfall's Completionist Log.xlsx"

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")


def import_quests(wb) -> list[dict]:
    """Extract Quest Journal sheet."""
    sheet = wb["Quest Journal"]
    quests = []
    # Headers are at row 2
    headers = [sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)]
    for row_idx in range(3, sheet.max_row + 1):
        row = [sheet.cell(row=row_idx, column=c).value for c in range(1, sheet.max_column + 1)]
        if row[0] is None:
            continue
        entry = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            v = row[i] if i < len(row) else None
            if v is None:
                continue
            # Clean up keys for JSON
            key = (h.lower()
                     .replace(" ", "_")
                     .replace("/", "_")
                     .replace("-", "_"))
            entry[key] = v
        quests.append(entry)
    return quests


def import_items(wb) -> list[dict]:
    """Extract Collection Log sheet."""
    sheet = wb["Collection Log"]
    items = []
    headers = [sheet.cell(row=2, column=c).value for c in range(1, sheet.max_column + 1)]
    for row_idx in range(3, sheet.max_row + 1):
        row = [sheet.cell(row=row_idx, column=c).value for c in range(1, sheet.max_column + 1)]
        if row[0] is None:
            continue
        entry = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            v = row[i] if i < len(row) else None
            if v is None:
                continue
            key = (h.lower()
                     .replace(" ", "_")
                     .replace("/", "_")
                     .replace("-", "_"))
            entry[key] = v
        items.append(entry)
    return items


def import_fragments(wb) -> dict:
    """Extract Death's Exchange (Fragments → item) costs."""
    sheet = wb["Fragments Exchange"]
    fragments = {}  # item_id → soul_fragments_required
    for row_idx in range(3, sheet.max_row + 1):
        item_id = sheet.cell(row=row_idx, column=1).value
        cost = sheet.cell(row=row_idx, column=3).value
        if item_id and cost is not None:
            fragments[str(item_id)] = float(cost)
    return fragments


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx_path):
        print(f"ERROR: XLSX not found at {xlsx_path}")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    quests = import_quests(wb)
    items = import_items(wb)
    fragments = import_fragments(wb)

    # Merge fragment cost into items for convenience
    for item in items:
        iid = item.get("id")
        if iid and iid in fragments:
            item["soul_fragments_cost"] = fragments[iid]

    os.makedirs(DATA_DIR, exist_ok=True)

    quests_path = os.path.join(DATA_DIR, "quests.json")
    items_path = os.path.join(DATA_DIR, "items.json")
    meta_path = os.path.join(DATA_DIR, "catalog_meta.json")

    with open(quests_path, "w") as f:
        json.dump(quests, f, indent=2, default=str)
    with open(items_path, "w") as f:
        json.dump(items, f, indent=2, default=str)

    meta = {
        "source": "Ashenfall's Completionist Log by Accurious (RSDW community)",
        "source_xlsx": os.path.basename(xlsx_path),
        "imported_at": datetime.now().isoformat(),
        "quest_count": len(quests),
        "item_count": len(items),
        "note": "Catalog uses community IDs (MQ01, L001). Game GUIDs not yet mapped.",
    }
    with open(meta_path, "w") as f:
        json.dump(meta, f, indent=2)

    print(f"\nWrote {len(quests)} quests → {quests_path}")
    print(f"Wrote {len(items)} items → {items_path}")
    print(f"Wrote meta → {meta_path}")

    # Print summary
    print("\nQuest types:", sorted(set(q.get("quest_type", "") for q in quests if q.get("quest_type"))))
    print("Item types:", sorted(set(i.get("type", "") for i in items if i.get("type"))))


if __name__ == "__main__":
    main()
